"""Export the current model results into one reviewable Excel workbook.

The export is deliberately read-only with respect to experiments: it only
combines existing node, OD, routing, and calibration result tables.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
EXPERIMENT_DIR = SCRIPT_DIR.parent
DEFAULT_OUTPUTS = EXPERIMENT_DIR / "outputs_10seed"


def gate05_name(value: object) -> object:
    """Give legacy thresholded calibration outputs an unambiguous name."""
    if not isinstance(value, str):
        return value
    replacements = {
        "Stable-Tail-EdgeTail-": "Stable-Tail-EdgeTailGate05-",
        "Stable-Tail-TailOnly-": "Stable-Tail-TailOnlyGate05-",
        "Stable-Tail-NodeCalib-": "Stable-Tail-NodeCalibGate05-",
        "Stable-Tail-MatrixEns-": "Stable-Tail-MatrixEnsGate05-",
    }
    for old, new in replacements.items():
        if value.startswith(old) and "Gate05" not in value:
            return value.replace(old, new, 1)
    return value


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Required result table is missing: {path}")
    return pd.read_csv(path)


def add_mode(frame: pd.DataFrame, mode: str, legacy_gate: bool = False) -> pd.DataFrame:
    result = frame.copy()
    result.insert(0, "result_family", mode)
    if legacy_gate:
        for column in ("model", "risk_source", "label"):
            if column in result:
                result[column] = result[column].map(gate05_name)
    return result


def make_model_master(
    node: pd.DataFrame,
    base_ab: pd.DataFrame,
    raw_ab: pd.DataFrame,
    gate_ab: pd.DataFrame,
    od_base: pd.DataFrame,
    od_raw: pd.DataFrame,
    od_gate: pd.DataFrame,
) -> pd.DataFrame:
    node_columns = [
        "model",
        "runs",
        "macro_f1_mean",
        "weighted_f1_mean",
        "mae_mean",
        "qwk_mean",
        "recall_6_8_mean",
        "precision_6_8_mean",
        "pr_auc_high_mean",
        "high_fn_rate_mean",
        "brier_mean",
        "ece_mean",
    ]
    node_master = node[node_columns].drop_duplicates("model")

    base_master = base_ab.copy()
    base_master.insert(1, "model_family", "trained_base_model")
    raw_master = raw_ab.copy()
    raw_master.insert(1, "model_family", "raw_p_high_calibration")
    gate_master = gate_ab.copy()
    gate_master.insert(1, "model_family", "gate05_calibration")
    budget_master = pd.concat([base_master, raw_master, gate_master], ignore_index=True)

    od_base = od_base.loc[od_base["method"].eq("CVaR-risk")].copy()
    od_raw = od_raw.loc[od_raw["method"].eq("CVaR-risk")].copy()
    od_gate = od_gate.loc[od_gate["method"].eq("CVaR-risk")].copy()
    od_gate["model"] = od_gate["model"].map(gate05_name)
    od = pd.concat([od_base, od_raw, od_gate], ignore_index=True)
    od_columns = [
        "model",
        "hop_increase_mean",
        "total_risk_reduction_mean",
        "cvar90_reduction_mean",
        "maxrisk_reduction_mean",
        "re_reduction_mean",
    ]
    od = od[od_columns].drop_duplicates("model")
    od = od.rename(
        columns={
            "hop_increase_mean": "od_hop_increase_mean",
            "total_risk_reduction_mean": "od_total_risk_reduction_mean",
            "cvar90_reduction_mean": "od_cvar90_reduction_mean",
            "maxrisk_reduction_mean": "od_maxrisk_reduction_mean",
            "re_reduction_mean": "od_re_reduction_mean",
        }
    )
    master = budget_master.merge(node_master, on="model", how="outer")
    master = master.merge(od, on="model", how="outer")
    master["model_family"] = master["model_family"].fillna("trained_base_model")
    return master.sort_values(["model_family", "model"], kind="stable").reset_index(drop=True)


def write_sheet(writer: pd.ExcelWriter, name: str, frame: pd.DataFrame) -> None:
    frame.to_excel(writer, sheet_name=name[:31], index=False)
    worksheet = writer.sheets[name[:31]]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    # The large detail sheets remain usable without inflating the file size.
    for index, column in enumerate(frame.columns, start=1):
        sample = [str(column)] + [str(value) for value in frame[column].head(100).fillna("")]
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(value) for value in sample) + 2, 36
        )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--outputs-root", type=Path, default=DEFAULT_OUTPUTS)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUTS / "all_models_complete_export")
    args = parser.parse_args()
    root = args.outputs_root
    out = args.output_dir
    out.mkdir(parents=True, exist_ok=True)

    calibration_root = root / "pyvrp_cvrp" / "stable_tail_calibration_raw_budget_sweep_10seed"
    legacy_calibration_root = root / "pyvrp_cvrp" / "stable_tail_calibration_budget_sweep_10seed"

    node = read_csv(root / "all_model_results_current" / "all_model_node_test_results.csv")
    od_base = read_csv(root / "all_model_results_current" / "all_model_od_results.csv")
    od_raw = read_csv(root / "od_paths" / "stable_tail_calibration_raw_od_10seed" / "od_seed_robustness_by_model_10seed.csv")
    od_gate = read_csv(root / "all_model_results_current" / "stable_tail_calibration_od_results_v3.csv")
    budget_base = read_csv(root / "pyvrp_cvrp" / "paper_budget_sweep_10seed" / "budget_sweep_summary.csv")
    budget_base_ab = read_csv(root / "pyvrp_cvrp" / "paper_budget_sweep_10seed" / "budget_sweep_ab_average.csv")
    raw_summary = read_csv(calibration_root / "raw_calibration_summary.csv")
    gate_summary = read_csv(calibration_root / "gate05_calibration_summary.csv")
    raw_ab = read_csv(calibration_root / "raw_calibration_ab_average.csv")
    gate_ab = read_csv(calibration_root / "gate05_calibration_ab_average.csv")
    pair = read_csv(calibration_root / "raw_vs_gate05_comparison.csv")
    raw_manifest = read_csv(root / "risk_matrices" / "stable_tail_calibration_raw_sources_10seed.csv")
    gate_manifest = read_csv(root / "risk_matrices" / "stable_tail_calibration_sources_10seed.csv")

    raw_pyvrp = read_csv(root / "pyvrp_cvrp" / "stable_tail_calibration_raw_pyvrp_10seed" / "model_pyvrp_summary.csv")
    gate_pyvrp = read_csv(root / "pyvrp_cvrp" / "stable_tail_calibration_pyvrp_10seed" / "model_pyvrp_summary.csv")
    raw_common = read_csv(root / "pyvrp_cvrp" / "stable_tail_calibration_raw_common_eval_10seed" / "common_route_summary.csv")
    gate_common = read_csv(root / "pyvrp_cvrp" / "stable_tail_calibration_common_eval_10seed" / "common_route_summary.csv")
    raw_load = read_csv(root / "pyvrp_cvrp" / "stable_tail_calibration_raw_load_eval_10seed" / "load_aware_summary.csv")
    gate_load = read_csv(root / "pyvrp_cvrp" / "stable_tail_calibration_load_eval_10seed" / "load_aware_summary.csv")
    raw_detail = read_csv(calibration_root / "budget_sweep_detail.csv")
    gate_detail = read_csv(legacy_calibration_root / "budget_sweep_detail.csv")

    master = make_model_master(node, budget_base_ab, raw_ab, gate_ab, od_base, od_raw, od_gate)
    sheets: list[tuple[str, pd.DataFrame]] = [
        (
            "README",
            pd.DataFrame(
                [
                    ("Export date", "2026-07-21"),
                    ("Comparable models", "47 = 13 trained models + 17 Raw calibrations + 17 Gate05 calibrations"),
                    ("Node metrics", "Only trained models; calibrations are post-processing and therefore intentionally blank."),
                    ("Metric direction", "All downstream budget metrics are lower-is-better. OD reductions are higher-is-better; OD hop increase is a cost."),
                    ("Raw semantics", "Raw uses P_high without the 0.5 threshold. Gate05 retains the legacy P_high >= 0.5 threshold."),
                    ("Selection caveat", "The best calibration is selected on the same A/B evaluation set and is exploratory, not independently confirmed."),
                    ("Detail sheets", "Rows retain seed, beta, customer set, and route-evaluation results where available."),
                ],
                columns=["topic", "detail"],
            ),
        ),
        ("Model_Master", master),
        ("Node_Test_13", node),
        ("OD_All", pd.concat([add_mode(od_base, "trained_base_model"), add_mode(od_raw, "raw_p_high"), add_mode(od_gate, "gate05", True)], ignore_index=True)),
        ("Budget_by_set", pd.concat([add_mode(budget_base, "trained_base_model"), add_mode(raw_summary, "raw_p_high"), add_mode(gate_summary, "gate05")], ignore_index=True)),
        ("Budget_AB_47", pd.concat([add_mode(budget_base_ab, "trained_base_model"), add_mode(raw_ab, "raw_p_high"), add_mode(gate_ab, "gate05")], ignore_index=True)),
        ("Raw_vs_Gate05", pair),
        ("Raw_manifest_170", raw_manifest),
        ("Gate05_manifest_170", add_mode(gate_manifest, "gate05", True)),
        ("Raw_PyVRP_1700", add_mode(raw_pyvrp, "raw_p_high")),
        ("Gate05_PyVRP_1700", add_mode(gate_pyvrp, "gate05", True)),
        ("Raw_Common_1700", add_mode(raw_common, "raw_p_high")),
        ("Gate05_Common_1700", add_mode(gate_common, "gate05", True)),
        ("Raw_Load_1700", add_mode(raw_load, "raw_p_high")),
        ("Gate05_Load_1700", add_mode(gate_load, "gate05", True)),
        ("Raw_Budget_Detail", add_mode(raw_detail, "raw_p_high")),
        ("Gate05_Budget_Detail", add_mode(gate_detail, "gate05", True)),
    ]

    workbook_path = out / "all_models_all_results_20260721.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for name, frame in sheets:
            write_sheet(writer, name, frame)

    master_path = out / "all_models_model_level_summary_20260721.csv"
    master.to_csv(master_path, index=False, encoding="utf-8-sig")
    index_path = out / "README.md"
    lines = [
        "# 全模型结果导出索引",
        "",
        "- `all_models_all_results_20260721.xlsx`：完整可筛选工作簿。",
        "- `all_models_model_level_summary_20260721.csv`：47 个可比较模型的模型级汇总。",
        "",
        "| 表 | 行数 |",
        "|---|---:|",
    ]
    lines.extend(f"| {name} | {len(frame)} |" for name, frame in sheets if name != "README")
    lines.extend(
        [
            "",
            "Raw 使用未阈值化 P_high；Gate05 为历史 P_high >= 0.5 对照。",
            "校准方案为风险矩阵后处理，故节点分类列对校准行留空。",
        ]
    )
    index_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {workbook_path}")
    print(f"Wrote {master_path}")
    print(f"Wrote {index_path}")


if __name__ == "__main__":
    main()
